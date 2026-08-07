"""Document services for tabular files.

This module provides the shared parsing pipeline that converts CSV and Excel
files directly to Markdown.
"""

from __future__ import annotations

import csv
import logging
from io import StringIO
from pathlib import Path
from typing import Any, Optional

try:
    import pandas as pd
    import chardet
    from openpyxl import load_workbook
    import xlrd
except ImportError:
    pd = None
    chardet = None
    load_workbook = None
    xlrd = None

from .simple_document_service import SimpleDocumentService

logger = logging.getLogger(__name__)


class CsvContentExtractor:
    """Extract content from CSV files."""

    _DELIMITER_CANDIDATES = (",", "\t", ";", "|", ":")
    _DELIMITER_SAMPLE_SIZE = 65536
    _DELIMITER_SAMPLE_ROWS = 200

    async def extract_content(self, file_path: Path) -> tuple[str, dict[str, Any]]:
        if pd is None or chardet is None:
            raise RuntimeError("未安装必需的库。请安装: pip install pandas chardet")

        try:
            df, meta = await self._read_csv_with_fallback(file_path)
            meta.update(
                {
                    "provider": "pandas",
                    "sheet_count": 1,
                    "row_count": len(df.index),
                    "column_count": len(df.columns),
                    "cell_count": int(df.shape[0] * df.shape[1]),
                    "non_empty_cell_count": int(df.notna().sum().sum()),
                }
            )
            try:
                return df.to_markdown(index=False, tablefmt="github"), meta
            except BaseException:
                logger.warning(
                    "csv_content_extractor failed to convert to markdown, using text format | file_path=%s",
                    file_path,
                    exc_info=True,
                )
                return df.to_string(), meta
        except BaseException as exc:
            logger.warning(
                "csv_content_extractor.extract_content failed | file_path=%s error=%s",
                file_path,
                exc,
                exc_info=True,
            )
            raise RuntimeError(f"提取CSV内容失败: {exc}") from exc

    async def _read_csv_with_fallback(self, file_path: Path):
        encodings = await self._get_encoding_candidates(file_path)
        last_error: Optional[BaseException] = None

        for encoding in encodings:
            try:
                delimiter, delimiter_detection = await self._detect_delimiter(file_path, encoding)
                df, parser_engine, fallback_reason = self._read_dataframe(
                    file_path=file_path,
                    encoding=encoding,
                    delimiter=delimiter,
                )
                return df, {
                    "encoding": encoding,
                    "delimiter": delimiter,
                    "delimiter_detection": delimiter_detection,
                    "parser_engine": parser_engine,
                    "fallback_reason": fallback_reason,
                }
            except (UnicodeDecodeError, LookupError) as exc:
                last_error = exc
                continue
            except BaseException as exc:
                last_error = exc
                break

        if isinstance(last_error, (UnicodeDecodeError, LookupError)):
            for encoding in encodings:
                try:
                    content = file_path.read_text(encoding=encoding, errors="replace")
                    delimiter, delimiter_detection = self._detect_delimiter_from_sample(content)
                    df, parser_engine, fallback_reason = self._read_dataframe(
                        content=content,
                        delimiter=delimiter,
                    )
                    fallback_parts = [part for part in ("decode_errors_replaced", fallback_reason) if part]
                    return df, {
                        "encoding": encoding,
                        "encoding_errors": "replace",
                        "delimiter": delimiter,
                        "delimiter_detection": delimiter_detection,
                        "parser_engine": parser_engine,
                        "fallback_reason": ",".join(fallback_parts),
                    }
                except BaseException as exc:
                    last_error = exc
                    continue

        if last_error is not None:
            raise last_error
        raise RuntimeError("无法识别CSV编码")

    async def _get_encoding_candidates(self, file_path: Path) -> list[str]:
        try:
            raw_data = file_path.read_bytes()[:65536]
            if raw_data.startswith(b"\xef\xbb\xbf"):
                detected = "utf-8-sig"
            elif raw_data.startswith(b"\xff\xfe") or raw_data.startswith(b"\xfe\xff"):
                detected = "utf-16"
            else:
                result = chardet.detect(raw_data)
                detected = result.get("encoding")
                confidence = result.get("confidence", 0)
                if not detected or confidence <= 0.7:
                    detected = "utf-8"

            candidates = [
                detected,
                "utf-8-sig",
                "utf-8",
                "utf-16",
                "utf-16-le",
                "utf-16-be",
                "gb18030",
                "gbk",
                "gb2312",
                "latin-1",
            ]
            deduped = []
            for encoding in candidates:
                if encoding and encoding not in deduped:
                    deduped.append(encoding)
            return deduped
        except BaseException:
            logger.warning(
                "csv_content_extractor._get_encoding_candidates failed, using defaults | file_path=%s",
                file_path,
                exc_info=True,
            )
            return ["utf-8", "utf-8-sig", "utf-16", "gb18030", "gbk", "latin-1"]

    @staticmethod
    def _read_dataframe(
        *,
        delimiter: str,
        file_path: Optional[Path] = None,
        encoding: Optional[str] = None,
        content: Optional[str] = None,
    ):
        source = file_path if file_path is not None else StringIO(content or "")
        kwargs: dict[str, Any] = {
            "delimiter": delimiter,
            "low_memory": False,
        }
        if file_path is not None:
            kwargs["encoding"] = encoding
        try:
            return pd.read_csv(source, engine="c", **kwargs), "c", ""
        except pd.errors.ParserError as c_error:
            if content is None and file_path is not None:
                source = file_path
            else:
                source = StringIO(content or "")
            python_kwargs = dict(kwargs)
            python_kwargs.pop("low_memory", None)
            try:
                return pd.read_csv(source, engine="python", **python_kwargs), "python", "c_parser_error"
            except BaseException as python_error:
                raise RuntimeError(
                    "CSV dialect unsupported after C and Python parser attempts: "
                    f"c={c_error}; python={python_error}"
                ) from python_error

    async def _detect_delimiter(self, file_path: Path, encoding: str) -> tuple[str, str]:
        try:
            sample = file_path.read_text(encoding=encoding)[: self._DELIMITER_SAMPLE_SIZE]
            return self._detect_delimiter_from_sample(sample)
        except (UnicodeDecodeError, LookupError):
            raise
        except BaseException:
            logger.warning(
                "csv_content_extractor._detect_delimiter failed, using comma | file_path=%s encoding=%s",
                file_path,
                encoding,
                exc_info=True,
            )
            return ",", "default"

    @classmethod
    def _detect_delimiter_from_sample(cls, sample: str) -> tuple[str, str]:
        complete_sample = sample
        if sample and not sample.endswith(("\n", "\r")):
            last_newline = max(sample.rfind("\n"), sample.rfind("\r"))
            if last_newline >= 0:
                complete_sample = sample[: last_newline + 1]
        if not complete_sample.strip():
            return ",", "default"

        sniffed: Optional[str] = None
        try:
            sniffed = csv.Sniffer().sniff(
                complete_sample,
                delimiters="".join(cls._DELIMITER_CANDIDATES),
            ).delimiter
        except csv.Error:
            pass

        candidates = list(dict.fromkeys([sniffed, *cls._DELIMITER_CANDIDATES]))
        candidates = [candidate for candidate in candidates if candidate]
        delimiter = max(
            candidates,
            key=lambda candidate: (
                *cls._delimiter_consistency_score(complete_sample, candidate),
                candidate == sniffed,
            ),
        )
        detection = "csv_sniffer_validated" if delimiter == sniffed else "column_consistency"
        return delimiter, detection

    @classmethod
    def _delimiter_consistency_score(cls, sample: str, delimiter: str) -> tuple[int, float, int, int]:
        rows = []
        try:
            for row in csv.reader(StringIO(sample), delimiter=delimiter):
                if row and any(cell.strip() for cell in row):
                    rows.append(row)
                if len(rows) >= cls._DELIMITER_SAMPLE_ROWS:
                    break
        except csv.Error:
            return 0, 0.0, 0, 0
        if not rows:
            return 0, 0.0, 0, 0

        header_width = len(rows[0])
        body_widths = [len(row) for row in rows[1:]]
        matching_rows = sum(width == header_width for width in body_widths)
        consistency = matching_rows / len(body_widths) if body_widths else 1.0
        return int(header_width > 1), consistency, matching_rows, header_width


class ExcelContentExtractor:
    """Extract content from Excel files."""

    async def extract_content(self, file_path: Path) -> tuple[str, dict[str, str | int]]:
        if pd is None:
            raise RuntimeError("未安装pandas。请安装: pip install pandas")

        if file_path.suffix.lower() == ".xls":
            if xlrd is None:
                raise RuntimeError("解析xls需要xlrd。请安装: pip install xlrd")
            markdown = await self._extract_xls_content(file_path)
            return markdown, self._inspect_xls(file_path)

        if load_workbook is None:
            raise RuntimeError("解析xlsx需要openpyxl。请安装: pip install openpyxl")
        markdown = await self._extract_xlsx_content(file_path)
        return markdown, self._inspect_xlsx(file_path)

    @staticmethod
    def _inspect_xlsx(file_path: Path) -> dict[str, str | int]:
        workbook = load_workbook(file_path, data_only=False, read_only=False)
        worksheets = list(workbook.worksheets)
        metrics: dict[str, str | int] = {
            "provider": "openpyxl",
            "file_format": "xlsx",
            "sheet_count": len(worksheets),
            "row_count": sum(sheet.max_row for sheet in worksheets),
            "column_count": max((sheet.max_column for sheet in worksheets), default=0),
            "cell_count": sum(sheet.max_row * sheet.max_column for sheet in worksheets),
            "non_empty_cell_count": sum(
                1 for sheet in worksheets for row in sheet.iter_rows() for cell in row if cell.value is not None
            ),
            "formula_count": sum(
                1
                for sheet in worksheets
                for row in sheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ),
            "merged_cell_count": sum(len(sheet.merged_cells.ranges) for sheet in worksheets),
            "chart_count": sum(len(getattr(sheet, "_charts", [])) for sheet in worksheets),
            "hidden_sheet_count": sum(1 for sheet in worksheets if sheet.sheet_state != "visible"),
        }
        workbook.close()
        return metrics

    @staticmethod
    def _inspect_xls(file_path: Path) -> dict[str, str | int]:
        workbook = xlrd.open_workbook(str(file_path), on_demand=True)
        sheets = [workbook.sheet_by_index(index) for index in range(workbook.nsheets)]
        metrics = {
            "provider": "xlrd",
            "file_format": "xls",
            "sheet_count": len(sheets),
            "row_count": sum(sheet.nrows for sheet in sheets),
            "column_count": max((sheet.ncols for sheet in sheets), default=0),
            "cell_count": sum(sheet.nrows * sheet.ncols for sheet in sheets),
            "non_empty_cell_count": sum(
                1 for sheet in sheets for row in range(sheet.nrows) for value in sheet.row_values(row) if value != ""
            ),
        }
        workbook.release_resources()
        return metrics

    async def _extract_xlsx_content(self, file_path: Path) -> str:
        try:
            workbook = load_workbook(file_path, data_only=True)
            content_parts = []
            content_parts.append(f"# {file_path.stem}\n\n")
            content_parts.append(f"**工作表数**: {len(workbook.sheetnames)}\n\n")
            content_parts.append("---\n\n")

            for sheet_name in workbook.sheetnames:
                content_parts.append(f"## 工作表: {sheet_name}\n\n")
                try:
                    worksheet = workbook[sheet_name]
                    if worksheet.max_row == 0 or worksheet.max_column == 0:
                        content_parts.append("*工作表为空*\n\n")
                        continue

                    merged_cells = await self._process_merged_cells(worksheet)
                    data = []
                    for row_idx, row in enumerate(
                        worksheet.iter_rows(
                            min_row=1,
                            max_row=worksheet.max_row,
                            min_col=1,
                            max_col=worksheet.max_column,
                            values_only=True,
                        ),
                        start=0,
                    ):
                        row_data = []
                        for col_idx, cell_value in enumerate(row):
                            if (row_idx, col_idx) in merged_cells:
                                merged_value = merged_cells[(row_idx, col_idx)]
                                cell_value = "" if merged_value is None else merged_value
                            if cell_value is None or (isinstance(cell_value, float) and pd.isna(cell_value)):
                                cell_value = ""
                            else:
                                cell_value = str(cell_value).strip()
                            row_data.append(cell_value)
                        data.append(row_data)

                    if not data:
                        content_parts.append("*工作表为空*\n\n")
                        continue

                    df = pd.DataFrame(data)
                    df = df.replace("", None).dropna(how="all").dropna(axis=1, how="all").fillna("")
                    if df.empty:
                        content_parts.append("*工作表为空*\n\n")
                        continue

                    markdown_lines = []
                    headers = [str(col) for col in df.columns]
                    markdown_lines.append("| " + " | ".join(headers) + " |")
                    markdown_lines.append("|" + "---|" * len(headers))
                    for _, row in df.iterrows():
                        row_values = [str(val) if val != "" else "" for val in row]
                        markdown_lines.append("| " + " | ".join(row_values) + " |")
                    content_parts.append("\n".join(markdown_lines) + "\n\n")
                except BaseException as exc:
                    logger.warning(
                        "excel_content_extractor failed to read worksheet | file_path=%s sheet_name=%s error=%s",
                        file_path,
                        sheet_name,
                        exc,
                        exc_info=True,
                    )
                    content_parts.append(f"**错误**: 无法读取工作表 - {exc}\n\n")

            workbook.close()
            return "".join(content_parts)
        except BaseException as exc:
            logger.warning(
                "excel_content_extractor._extract_xlsx_content failed | file_path=%s error=%s",
                file_path,
                exc,
                exc_info=True,
            )
            raise RuntimeError(f"提取XLSX内容失败: {exc}") from exc

    async def _extract_xls_content(self, file_path: Path) -> str:
        try:
            excel_file = pd.ExcelFile(file_path, engine="xlrd")
            content_parts = []
            content_parts.append(f"# {file_path.stem}\n\n")
            content_parts.append(f"**工作表数**: {len(excel_file.sheet_names)}\n\n")
            content_parts.append("---\n\n")

            for sheet_name in excel_file.sheet_names:
                content_parts.append(f"## 工作表: {sheet_name}\n\n")
                try:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        engine="xlrd",
                        header=None,
                    )
                    df = df.dropna(how="all").dropna(axis=1, how="all")
                    if df.empty:
                        content_parts.append("*工作表为空*\n\n")
                        continue
                    df = df.fillna("")
                    headers = [str(col) for col in df.columns]
                    markdown_lines = []
                    markdown_lines.append("| " + " | ".join(headers) + " |")
                    markdown_lines.append("|" + "---|" * len(headers))
                    for _, row in df.iterrows():
                        row_values = [str(val).strip() if val != "" else "" for val in row]
                        markdown_lines.append("| " + " | ".join(row_values) + " |")
                    content_parts.append("\n".join(markdown_lines) + "\n\n")
                except BaseException as exc:
                    logger.warning(
                        "excel_content_extractor failed to read xls worksheet | file_path=%s sheet_name=%s error=%s",
                        file_path,
                        sheet_name,
                        exc,
                        exc_info=True,
                    )
                    content_parts.append(f"**错误**: 无法读取工作表 - {exc}\n\n")

            excel_file.close()
            return "".join(content_parts)
        except BaseException as exc:
            logger.warning(
                "excel_content_extractor._extract_xls_content failed | file_path=%s error=%s",
                file_path,
                exc,
                exc_info=True,
            )
            raise RuntimeError(f"提取XLS内容失败: {exc}") from exc

    async def _process_merged_cells(self, worksheet):
        merged_cells_map = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col_idx, max_row_idx = merged_range.bounds
            min_col -= 1
            min_row -= 1
            max_col_idx -= 1
            max_row_idx -= 1

            top_left_value = worksheet.cell(row=min_row + 1, column=min_col + 1).value
            for row_idx in range(min_row, max_row_idx + 1):
                for col_idx in range(min_col, max_col_idx + 1):
                    if row_idx == min_row and col_idx == min_col:
                        merged_cells_map[(row_idx, col_idx)] = top_left_value
                    else:
                        merged_cells_map[(row_idx, col_idx)] = None
        return merged_cells_map


class CsvDocumentService(SimpleDocumentService):
    """Top-level document service for CSV files."""

    def __init__(self) -> None:
        super().__init__(file_type="csv", content_extractor=CsvContentExtractor())


class ExcelDocumentService(SimpleDocumentService):
    """Top-level document service for Excel files."""

    def __init__(self) -> None:
        super().__init__(file_type="xlsx", content_extractor=ExcelContentExtractor())
